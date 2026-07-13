"""
O módulo excel_exporter fornece utilitários para exportação e formatação visual
dos resultados obtidos nas rodadas do benchmark para planilhas do Microsoft Excel.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Final
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

@dataclass(slots=True, frozen=True)
class BenchmarkResult:
    """
    Representa a estrutura de dados consolidada com as métricas obtidas por um modelo em uma rodada.
    """
    arquivo: str
    eventos: int
    tempo_total: float | str
    throughput: float | str
    status: str

def export_results_to_excel(results: list[BenchmarkResult], filename: str) -> None:
    """
    Agrupa, formata e exporta os resultados do benchmark para uma planilha Excel.

    A planilha gerada organiza cada rodada de benchmark (com base no número de eventos)
    em colunas consecutivas e estilizadas com fontes personalizadas, cabeçalhos destacados,
    bordas e redimensionamento automático de colunas para garantir legibilidade.

    Args:
        results (list[BenchmarkResult]): Lista de resultados consolidados das execuções do benchmark.
        filename (str): O caminho do arquivo Excel (.xlsx) de saída a ser salvo.
    """
    workbook: Final[Workbook] = Workbook()
    sheet: Final[Worksheet] = workbook.active
    sheet.title = "Benchmark Results"
    sheet.views.sheetView[0].showGridLines = True

    rounds: Final[dict[int, list[BenchmarkResult]]] = {}
    rounds_keys: Final[list[int]] = []
    
    for result in results:
        if result.status == "Pulado":
            continue
        if result.eventos not in rounds:
            rounds[result.eventos] = []
            rounds_keys.append(result.eventos)
        rounds[result.eventos].append(result)

    title_font: Final[Font] = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
    header_font: Final[Font] = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font: Final[Font] = Font(name="Segoe UI", size=10)
    
    header_fill: Final[PatternFill] = PatternFill(
        start_color="2F5597", end_color="2F5597", fill_type="solid"
    )
    
    thin_border_side: Final[Side] = Side(border_style="thin", color="D9D9D9")
    border_style: Final[Border] = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )
    
    alignment_center: Final[Alignment] = Alignment(horizontal="center", vertical="center")
    alignment_left: Final[Alignment] = Alignment(horizontal="left", vertical="center")

    for index in range(len(rounds_keys)):
        event_count: Final[int] = rounds_keys[index]
        round_results: Final[list[BenchmarkResult]] = rounds[event_count]
        round_index: Final[int] = index + 1
        
        start_col: Final[int] = (round_index - 1) * 5 + 1

        sheet.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=start_col + 3
        )
        title_cell: Final[Cell] = sheet.cell(
            row=1, column=start_col, value=f"Rodada {round_index} - {event_count:,} Eventos"
        )
        title_cell.font = title_font
        title_cell.alignment = alignment_center

        headers: Final[list[str]] = ["Arquivo", "Tempo total (s)", "Throughput (ev/s)", "Status"]
        for col_idx in range(len(headers)):
            header: Final[str] = headers[col_idx]
            cell: Final[Cell] = sheet.cell(row=2, column=start_col + col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = border_style

        for row_idx in range(len(round_results)):
            result: Final[BenchmarkResult] = round_results[row_idx]
            row_data: Final[list[str | float]] = [
                result.arquivo,
                result.tempo_total,
                result.throughput,
                result.status
            ]
            for col_idx in range(len(row_data)):
                val: Final[str | float] = row_data[col_idx]
                cell: Final[Cell] = sheet.cell(
                    row=row_idx + 3, column=start_col + col_idx, value=val
                )
                cell.font = data_font
                cell.border = border_style
                if col_idx == 0:
                    cell.alignment = alignment_left
                else:
                    cell.alignment = alignment_center

    for col_cells in sheet.columns:
        max_length: int = 0
        first_cell: Final[Cell] = col_cells[0]
        column_letter: Final[str] = get_column_letter(first_cell.column)
        has_content: bool = False
        for cell in col_cells:
            cell_val: Final[object] = cell.value
            if cell_val is not None:
                has_content = True
                max_length = max(max_length, len(str(cell_val)))
        if has_content:
            sheet.column_dimensions[column_letter].width = max(max_length + 3, 12)
        else:
            sheet.column_dimensions[column_letter].width = 4

    workbook.save(filename)
